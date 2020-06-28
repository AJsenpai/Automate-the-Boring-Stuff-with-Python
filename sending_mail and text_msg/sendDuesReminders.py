# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.
import openpyxl, smtplib, sys

# open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook(r"C:\Users\Jai\Desktop\Docs\duesRecords.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# check each members payment status

unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# log in to email account

with smtplib.SMTP(host="smtp.gmail.com", port=587) as smtpObj:
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login("dj020199.lex@gmail.com", "*****")

    # set reminder emails
    for name, email in unpaidMembers.items():
        body = f"""Subject: {latestMonth} dues unpaid \n\nDear {name}, \nRecords shows that you have not paid dues for {latestMonth}. Please make
this payment as soon as possible.
Thank You! """
        print(f"sending mails to {email}...")
        sendmailStatus = smtpObj.sendmail("dj020199.lex@gmail.com", email, body)

        if sendmailStatus != {}:
            print(f"There was a problem sending email to {email}, {sendmailStatus}")
print("Done")
