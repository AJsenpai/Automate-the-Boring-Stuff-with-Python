# invert the cells in excel

import openpyxl

# print(f"c:/Users/Jai/Desktop/Docs/{sys.argv[1]}.xlsx")
# wb = openpyxl.load_workbook(f"c:/Users/Jai/Desktop/Docs/{sys.argv[1]}.xlsx")
wb = openpyxl.load_workbook(f"c:/Users/Jai/Desktop/Docs/pyexcel.xlsx")
ws = wb[wb.sheetnames[0]]
templist = []
for col in range(1, ws.max_column + 1):
    for row in range(ws.max_column + 1, ws.max_row + 1):
        newlist = ws.cell(row, col).value

templist.append(newlist)
print(templist)
wb.save((f"c:/Users/Jai/Desktop/Docs/pyexcel.xlsx"))
