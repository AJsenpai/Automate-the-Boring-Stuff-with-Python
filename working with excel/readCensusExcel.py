# readCensusExcel.py - Tabulates population and number of census tracts for each county.

import openpyxl, pprint

print("Opening Workbook....")

wb = openpyxl.load_workbook("c:/Users/Jai/Desktop/Docs/censuspopdata.xlsx")
# print(wb.sheetnames)
sheet = wb["Population by Census Tract"]
countyData = {}
# TODO : fill in county data with each population and tracts
print("reading rows...")
for row in range(2, sheet.max_row + 1):
    # Each row in the sheet has data for one census tract.
    state = sheet["B" + str(row)].value
    county = sheet["C" + str(row)].value
    pop = sheet["D" + str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {"tracts": 0, "pop": 0})

    # Each row represent one census tract, so increment by one
    countyData[state][county]["tracts"] += 1
    countyData[state][county]["pop"] += int(pop)

# TODO : open a new text file and write the contents of countdata to it
print("Writing results into a file")
with open(
    "Automate the Boring stuff with Python\working with excel\census2010.py", "w",
) as resultFile:
    resultFile.write("allData = " + pprint.pformat(countyData))
print("Done")
