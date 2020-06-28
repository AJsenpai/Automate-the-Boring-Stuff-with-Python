# updateProduce.py - Corrects costs in produce sales spreadsheet.
import openpyxl

wb = openpyxl.load_workbook("c:/Users/Jai/Desktop/Docs/produceSales.xlsx")
sheet = wb["Sheet"]
price_updates = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

for rowNum in range(2, sheet.max_row + 1):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_updates:
        sheet.cell(row=rowNum, column=2).value = price_updates[produceName]
wb.save("c:/Users/Jai/Desktop/Docs/produceSales_updated.xlsx")
